"""Generate an E-FIR report by matching acknowledgement numbers."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, Iterable, Mapping

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.ui_styling import render_page_header_with_info


REPORT_COLUMNS = [
    "Sr.",
    "ACK No.",
    "Name",
    "Fraudulent Amount",
    "District",
    "Police Station",
    "E-FIR No.",
]

CRIME_COLUMN_ALIASES = {
    "ACK No.": (
        "acknowledgementno",
        "acknowledgementnumber",
        "acknowledgmentno",
        "acknowledgmentnumber",
        "ackno",
        "acknumber",
        "ack",
    ),
    "Name": (
        "nameofcomplainant",
        "complainantname",
        "victimname",
        "name",
    ),
    "District": (
        "district",
        "complainantdistrict",
        "victimdistrict",
    ),
    "Police Station": (
        "policestation",
        "policestationname",
        "psname",
        "ps",
    ),
}

EFIR_COLUMN_ALIASES = {
    "ACK No.": CRIME_COLUMN_ALIASES["ACK No."],
    "E-FIR No.": (
        "ezerofirno",
        "ezerofirnumber",
        "efirno",
        "efirnumber",
        "zerofirno",
    ),
    "Fraudulent Amount": (
        "totalreportedamount",
        "reportedamount",
        "fraudulentamount",
        "fraudamount",
    ),
    "E-FIR Date": (
        "dateofezerofir",
        "dateofefir",
        "efirdate",
        "ezerofirdate",
    ),
}

_ILLEGAL_EXCEL_CHARS_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")
_NON_ALPHANUMERIC_RE = re.compile(r"[^A-Za-z0-9]+")
_AMOUNT_CLEAN_RE = re.compile(r"[^0-9.()\-]+")


@dataclass(frozen=True)
class EFIRReportResult:
    report: pd.DataFrame
    workbook_bytes: bytes
    filename: str
    summary: Dict[str, int]


def _normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value).lower())


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    text = str(value).strip()
    if text.lower() in {"nan", "none", "<na>", "nat"}:
        return ""
    return _ILLEGAL_EXCEL_CHARS_RE.sub("", text)


def _clean_identifier(value: Any) -> str:
    text = _clean_text(value)
    if not text:
        return ""

    if isinstance(value, str):
        if re.fullmatch(r"\d+\.0+", text):
            return text.split(".", 1)[0]
        if not re.fullmatch(r"\d+(?:\.\d+)?[eE][+-]?\d+", text):
            return text

    numeric_text = text.replace(",", "")
    try:
        number = Decimal(numeric_text)
        if number == number.to_integral_value():
            return format(number.quantize(Decimal("1")), "f")
    except (InvalidOperation, ValueError):
        pass
    return text


def _ack_key(value: Any) -> str:
    return _NON_ALPHANUMERIC_RE.sub("", _clean_identifier(value)).upper()


def _parse_amount(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            if pd.isna(value):
                return None
        except TypeError:
            pass
        return float(value)

    text = _clean_text(value)
    if not text:
        return None
    negative = text.startswith("(") and text.endswith(")")
    cleaned = _AMOUNT_CLEAN_RE.sub("", text).replace("(", "").replace(")", "")
    try:
        amount = float(Decimal(cleaned))
    except (InvalidOperation, ValueError):
        return None
    return -amount if negative else amount


def _parse_datetime(value: Any) -> datetime | None:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)

    text = _clean_text(value)
    if not text:
        return None
    parsed = pd.to_datetime(text, errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        return None
    return parsed.to_pydatetime()


def _first_non_empty(values: Iterable[Any]) -> str:
    for value in values:
        cleaned = _clean_text(value)
        if cleaned:
            return cleaned
    return ""


def _detect_columns(
    columns: Iterable[Any],
    aliases: Mapping[str, Iterable[str]],
    source_label: str,
) -> Dict[str, str]:
    normalized = {_normalize_header(column): str(column) for column in columns}
    detected: Dict[str, str] = {}
    for output_field, field_aliases in aliases.items():
        for alias in field_aliases:
            if alias in normalized:
                detected[output_field] = normalized[alias]
                break

    missing = [field for field in aliases if field not in detected]
    if missing:
        raise ValueError(
            f"{source_label} is missing required column(s): {', '.join(missing)}."
        )
    return detected


def _read_tabular_file(
    file_bytes: bytes,
    filename: str,
    preferred_sheet: str | None = None,
) -> pd.DataFrame:
    extension = Path(filename).suffix.lower()
    buffer = io.BytesIO(file_bytes)

    if extension == ".csv":
        last_error: Exception | None = None
        for encoding in ("utf-8-sig", "utf-8", "cp1252"):
            try:
                buffer.seek(0)
                return pd.read_csv(
                    buffer,
                    dtype=object,
                    keep_default_na=False,
                    encoding=encoding,
                )
            except UnicodeDecodeError as error:
                last_error = error
        raise ValueError(f"Unable to read {filename}: {last_error}")

    if extension in {".xlsx", ".xls"}:
        excel_file = pd.ExcelFile(buffer)
        sheet_name = excel_file.sheet_names[0]
        if preferred_sheet:
            preferred_key = _normalize_header(preferred_sheet)
            sheet_name = next(
                (
                    candidate
                    for candidate in excel_file.sheet_names
                    if _normalize_header(candidate) == preferred_key
                ),
                sheet_name,
            )
        return pd.read_excel(
            excel_file,
            sheet_name=sheet_name,
            dtype=object,
            keep_default_na=False,
        )

    raise ValueError(f"Unsupported file type for {filename}. Use CSV, XLSX, or XLS.")


def build_efir_report(
    crime_df: pd.DataFrame,
    efir_df: pd.DataFrame,
    report_date: date,
) -> tuple[pd.DataFrame, Dict[str, int]]:
    crime_columns = _detect_columns(
        crime_df.columns,
        CRIME_COLUMN_ALIASES,
        "Crime Report",
    )
    efir_columns = _detect_columns(
        efir_df.columns,
        EFIR_COLUMN_ALIASES,
        "Zero E-FIR Report",
    )

    crime_rows = pd.DataFrame(
        {
            "_ack_key": crime_df[crime_columns["ACK No."]].map(_ack_key),
            "ACK No.": crime_df[crime_columns["ACK No."]].map(_clean_identifier),
            "Name": crime_df[crime_columns["Name"]].map(_clean_text),
            "District": crime_df[crime_columns["District"]].map(_clean_text),
            "Police Station": crime_df[crime_columns["Police Station"]].map(
                _clean_text
            ),
        }
    )
    crime_rows = crime_rows[crime_rows["_ack_key"] != ""].copy()

    crime_grouped = (
        crime_rows.groupby("_ack_key", sort=False, as_index=False)
        .agg(
            {
                "ACK No.": _first_non_empty,
                "Name": _first_non_empty,
                "District": _first_non_empty,
                "Police Station": _first_non_empty,
            }
        )
        .reset_index(drop=True)
    )

    efir_rows = pd.DataFrame(
        {
            "_ack_key": efir_df[efir_columns["ACK No."]].map(_ack_key),
            "E-FIR No.": efir_df[efir_columns["E-FIR No."]].map(_clean_identifier),
            "Fraudulent Amount": efir_df[
                efir_columns["Fraudulent Amount"]
            ].map(_parse_amount),
            "_efir_datetime": efir_df[efir_columns["E-FIR Date"]].map(
                _parse_datetime
            ),
            "_efir_order": range(len(efir_df)),
        }
    )
    efir_rows["_efir_date"] = efir_rows["_efir_datetime"].map(
        lambda value: value.date() if value is not None else None
    )
    efir_rows = efir_rows[
        (efir_rows["_ack_key"] != "") & (efir_rows["_efir_date"] == report_date)
    ].copy()
    efir_grouped = (
        efir_rows.groupby("_ack_key", sort=False, as_index=False)
        .agg(
            {
                "E-FIR No.": _first_non_empty,
                "Fraudulent Amount": "first",
                "_efir_datetime": "min",
                "_efir_order": "min",
            }
        )
        .reset_index(drop=True)
    )

    matched = efir_grouped.merge(crime_grouped, on="_ack_key", how="inner", sort=False)
    matched = matched.sort_values(
        ["_efir_datetime", "_efir_order"],
        kind="stable",
    ).reset_index(drop=True)

    report = matched[
        [
            "ACK No.",
            "Name",
            "Fraudulent Amount",
            "District",
            "Police Station",
            "E-FIR No.",
        ]
    ].copy()
    report.insert(0, "Sr.", range(1, len(report) + 1))
    report = report[REPORT_COLUMNS]

    crime_ack_keys = set(crime_grouped["_ack_key"])
    efir_ack_keys = set(efir_grouped["_ack_key"])
    summary = {
        "crime_rows": len(crime_df),
        "crime_acknowledgements": len(crime_ack_keys),
        "efir_rows": len(efir_df),
        "efir_acknowledgements": len(efir_ack_keys),
        "selected_efir_rows": len(efir_rows),
        "matched_acknowledgements": len(report),
        "unmatched_crime_acknowledgements": len(crime_ack_keys - efir_ack_keys),
        "unmatched_efir_acknowledgements": len(efir_ack_keys - crime_ack_keys),
        "combined_crime_rows": len(crime_rows) - len(crime_grouped),
    }
    return report, summary


def generate_efir_excel(report: pd.DataFrame) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "E-FIR Report"
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"

    worksheet.append(REPORT_COLUMNS)
    for row in report.itertuples(index=False, name=None):
        worksheet.append([None if pd.isna(value) else value for value in row])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Calibri", size=10, color="172033")
    even_fill = PatternFill("solid", fgColor="EAF2F8")
    odd_fill = PatternFill("solid", fgColor="FFFFFF")
    border_side = Side(style="thin", color="CBD5E1")
    cell_border = Border(
        left=border_side,
        right=border_side,
        top=border_side,
        bottom=border_side,
    )
    centered = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_aligned = Alignment(horizontal="left", vertical="center", wrap_text=True)

    worksheet.row_dimensions[1].height = 28
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = cell_border
        cell.alignment = centered

    for row_number in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row_number].height = 24
        row_fill = even_fill if row_number % 2 == 0 else odd_fill
        for column_number in range(1, len(REPORT_COLUMNS) + 1):
            cell = worksheet.cell(row=row_number, column=column_number)
            cell.fill = row_fill
            cell.font = body_font
            cell.border = cell_border
            cell.alignment = centered if column_number in {1, 2, 4, 7} else left_aligned
        worksheet.cell(row=row_number, column=2).number_format = "@"
        worksheet.cell(row=row_number, column=4).number_format = "#,##0.00"
        worksheet.cell(row=row_number, column=7).number_format = "@"

    widths = {
        "A": 8,
        "B": 22,
        "C": 28,
        "D": 20,
        "E": 24,
        "F": 24,
        "G": 22,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    worksheet.auto_filter.ref = f"A1:G{max(worksheet.max_row, 1)}"
    if worksheet.max_row > 1:
        table = Table(displayName="EFIRReportTable", ref=f"A1:G{worksheet.max_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def process_efir_report_files(
    crime_file_bytes: bytes,
    crime_filename: str,
    efir_file_bytes: bytes,
    efir_filename: str,
    report_date: date,
) -> EFIRReportResult:
    crime_df = _read_tabular_file(crime_file_bytes, crime_filename)
    efir_df = _read_tabular_file(
        efir_file_bytes,
        efir_filename,
        preferred_sheet="Details",
    )
    report, summary = build_efir_report(crime_df, efir_df, report_date)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return EFIRReportResult(
        report=report,
        workbook_bytes=generate_efir_excel(report),
        filename=f"E_FIR_Report_{report_date:%Y%m%d}_{timestamp}.xlsx",
        summary=summary,
    )


def render_efir_report_page() -> None:
    render_page_header_with_info("e_fir_report")

    report_date = st.date_input(
        "E-FIR Date",
        value=date.today(),
        format="DD/MM/YYYY",
        key="efir_report_date",
    )

    upload_columns = st.columns(2)
    with upload_columns[0]:
        crime_file = st.file_uploader(
            "Crime Report",
            type=["csv", "xlsx", "xls"],
            key="efir_crime_report_upload",
        )
    with upload_columns[1]:
        efir_file = st.file_uploader(
            "Zero E-FIR Report",
            type=["xlsx", "xls"],
            key="efir_zero_report_upload",
        )

    if crime_file is None or efir_file is None:
        st.info("Upload both reports to generate the E-FIR Excel file.")
        return

    if not st.button(
        "Generate E-FIR Report",
        type="primary",
        use_container_width=True,
        key="generate_efir_report",
    ):
        return

    try:
        with st.spinner("Matching acknowledgement numbers..."):
            result = process_efir_report_files(
                crime_file.getvalue(),
                crime_file.name,
                efir_file.getvalue(),
                efir_file.name,
                report_date,
            )
    except (ValueError, KeyError, OSError) as error:
        st.error(str(error))
        return
    except Exception as error:
        st.error(f"Unable to generate the E-FIR report: {error}")
        return

    summary = result.summary
    metrics = st.columns(4)
    metrics[0].metric(
        "E-FIR ACKs on selected date",
        f"{summary['efir_acknowledgements']:,}",
    )
    metrics[1].metric(
        "Matched ACKs",
        f"{summary['matched_acknowledgements']:,}",
    )
    metrics[2].metric(
        "E-FIR ACKs not matched",
        f"{summary['unmatched_efir_acknowledgements']:,}",
    )
    metrics[3].metric(
        "Crime Report ACKs",
        f"{summary['crime_acknowledgements']:,}",
    )

    if result.report.empty:
        st.warning("No matching acknowledgement numbers were found.")
    else:
        st.subheader("Matched E-FIR Preview")
        st.dataframe(result.report, use_container_width=True, hide_index=True)

    st.download_button(
        "Download E-FIR Report Excel",
        data=result.workbook_bytes,
        file_name=result.filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
        disabled=result.report.empty,
        key="download_efir_report",
    )
