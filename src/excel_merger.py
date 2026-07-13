"""Smart Excel/CSV merger used by the Merge Excel Files page."""

from datetime import datetime
from io import BytesIO
import re
from typing import Dict, List, Sequence, Tuple

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


MAX_MERGE_FILES = 50

# Equivalent report headers share one output column while the first uploaded
# file still controls the visible column label.
HEADER_KEY_ALIASES = {
    "acknowledgementno": "ackno",
    "acknowledgementnumber": "ackno",
    "acknowledgmentno": "ackno",
    "acknowledgmentnumber": "ackno",
    "acknumber": "ackno",
    "bankfis": "bankname",
    "bankfi": "bankname",
    "bankfinancialinstitution": "bankname",
    "accountno": "accountnumber",
    "acno": "accountnumber",
    "accno": "accountnumber",
    "ifsccode": "ifsc",
    "transactionamount": "amount",
    "txnamount": "amount",
}

SOURCE_FILLS = [
    "F3F8FD",
    "F4FBF7",
    "FFF9ED",
    "F9F5FC",
    "FFF4F5",
    "F2FAFA",
]


def _is_blank(value) -> bool:
    if value is None or pd.isna(value):
        return True
    return isinstance(value, str) and not value.strip()


def _display_header(value, position: int) -> str:
    header = re.sub(r"\s+", " ", str(value).replace("\ufeff", "")).strip()
    if not header or header.lower().startswith("unnamed:"):
        return f"Column {position + 1}"
    return header


def _header_key(value) -> str:
    key = re.sub(r"[^a-z0-9]+", "", str(value).lower())
    return HEADER_KEY_ALIASES.get(key, key)


def _coalesce_columns(existing: pd.Series, incoming: pd.Series) -> pd.Series:
    result = existing.copy()
    empty_mask = result.map(_is_blank)
    result.loc[empty_mask] = incoming.loc[empty_mask]
    return result


def _remove_repeated_header_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    minimum_matches = max(1, (len(df.columns) + 1) // 2)

    def is_header_row(row: pd.Series) -> bool:
        populated = 0
        matches = 0
        for column_name, value in row.items():
            if _is_blank(value):
                continue
            populated += 1
            if _header_key(value) == _header_key(column_name):
                matches += 1
        return populated >= minimum_matches and matches == populated

    repeated_headers = df.apply(is_header_row, axis=1)
    return df.loc[~repeated_headers].reset_index(drop=True)


def _normalize_columns(
    df: pd.DataFrame,
    canonical_headers: Dict[str, str],
) -> pd.DataFrame:
    normalized = pd.DataFrame(index=df.index)

    for position, original_header in enumerate(df.columns):
        source = df.iloc[:, position]
        display_header = _display_header(original_header, position)

        # Ignore completely empty unnamed columns commonly created by Excel.
        if display_header.startswith("Column ") and source.map(_is_blank).all():
            continue

        key = _header_key(display_header)
        canonical = canonical_headers.setdefault(key, display_header)
        if canonical in normalized.columns:
            normalized[canonical] = _coalesce_columns(normalized[canonical], source)
        else:
            normalized[canonical] = source

    if normalized.empty and len(normalized.columns) == 0:
        return normalized

    blank_rows = normalized.apply(lambda column: column.map(_is_blank)).all(axis=1)
    normalized = normalized.loc[~blank_rows].reset_index(drop=True)
    return _remove_repeated_header_rows(normalized)


def _read_csv(raw_data: bytes) -> pd.DataFrame:
    last_error = None
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            frame = pd.read_csv(
                BytesIO(raw_data),
                encoding=encoding,
                dtype=str,
                keep_default_na=False,
                low_memory=False,
            )
        except UnicodeDecodeError as error:
            last_error = error
            continue
        except pd.errors.ParserError:
            frame = pd.read_csv(
                BytesIO(raw_data),
                encoding=encoding,
                dtype=str,
                keep_default_na=False,
                engine="python",
                on_bad_lines="skip",
            )

        # Handle common semicolon/tab/pipe-separated files saved as CSV.
        if len(frame.columns) == 1:
            first_line = raw_data.decode(encoding, errors="ignore").splitlines()[0]
            separator = next((item for item in (";", "\t", "|") if item in first_line), None)
            if separator:
                frame = pd.read_csv(
                    BytesIO(raw_data),
                    encoding=encoding,
                    sep=separator,
                    dtype=str,
                    keep_default_na=False,
                    engine="python",
                    on_bad_lines="skip",
                )
        return frame

    raise ValueError(f"Could not decode CSV file: {last_error}")


def read_file(uploaded_file) -> pd.DataFrame:
    """Read the first sheet from an uploaded Excel file or a CSV file."""
    filename = uploaded_file.name
    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    uploaded_file.seek(0)
    raw_data = uploaded_file.read()

    if not raw_data:
        raise ValueError(f"'{filename}' is empty.")

    try:
        if extension == "csv":
            return _read_csv(raw_data)
        if extension in {"xlsx", "xls"}:
            return pd.read_excel(
                BytesIO(raw_data),
                sheet_name=0,
                dtype=object,
                keep_default_na=False,
            )
    except Exception as error:
        raise ValueError(f"Could not read '{filename}': {error}") from error

    raise ValueError(f"Unsupported file type for '{filename}'.")


def merge_uploaded_files(
    uploaded_files: Sequence,
) -> Tuple[pd.DataFrame, List[Tuple[str, int, int]]]:
    """Stack uploaded files after normalizing equivalent and duplicate headers."""
    if not 2 <= len(uploaded_files) <= MAX_MERGE_FILES:
        raise ValueError(f"Select between 2 and {MAX_MERGE_FILES} files.")

    canonical_headers: Dict[str, str] = {}
    normalized_frames: List[pd.DataFrame] = []
    file_sources: List[Tuple[str, int, int]] = []
    current_row = 0

    for uploaded_file in uploaded_files:
        frame = _normalize_columns(read_file(uploaded_file), canonical_headers)
        if frame.empty or len(frame.columns) == 0:
            raise ValueError(f"'{uploaded_file.name}' has no usable data rows.")

        normalized_frames.append(frame)
        end_row = current_row + len(frame) - 1
        file_sources.append((uploaded_file.name, current_row, end_row))
        current_row = end_row + 1

    merged = pd.concat(normalized_frames, ignore_index=True, sort=False)
    ordered_columns = [
        column_name
        for column_name in canonical_headers.values()
        if column_name in merged.columns
    ]
    merged = merged.reindex(columns=ordered_columns).where(pd.notna(merged), "")

    serial_keys = {
        "sno",
        "serialno",
        "serialnumber",
        "srno",
        "slno",
    }
    for column_name in merged.columns:
        if _header_key(column_name) in serial_keys:
            merged[column_name] = range(1, len(merged) + 1)
            break

    return merged, file_sources


def _safe_excel_value(value):
    if isinstance(value, str) and value.startswith("="):
        return "'" + value
    return value


def generate_merged_excel(
    df: pd.DataFrame,
    file_sources: List[Tuple[str, int, int]] = None,
) -> bytes:
    """Create a polished Excel workbook with one merged header row."""
    if df.empty or len(df.columns) == 0:
        raise ValueError("There is no merged data to export.")

    file_sources = file_sources or [("Merged data", 0, len(df) - 1)]
    safe_df = df.map(_safe_excel_value)
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        safe_df.to_excel(writer, index=False, sheet_name="Merged Data")
        workbook = writer.book
        worksheet = writer.sheets["Merged Data"]

        last_column = get_column_letter(len(df.columns))
        last_row = len(df) + 1
        title_fill = PatternFill("solid", fgColor="132238")
        header_fill = PatternFill("solid", fgColor="087E8B")
        accent_fill = PatternFill("solid", fgColor="F2B134")
        white_font = Font(name="Aptos", color="FFFFFF", bold=True)
        light_border = Border(bottom=Side(style="hair", color="D9E2EC"))

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        worksheet.row_dimensions[1].height = 30

        for source_index, (_, start_row, end_row) in enumerate(file_sources):
            source_fill = PatternFill("solid", fgColor=SOURCE_FILLS[source_index % len(SOURCE_FILLS)])
            first_excel_row = start_row + 2
            last_excel_row = end_row + 2
            for row in worksheet.iter_rows(
                min_row=first_excel_row,
                max_row=last_excel_row,
                min_col=1,
                max_col=len(df.columns),
            ):
                for cell in row:
                    cell.fill = source_fill
                    cell.font = Font(name="Aptos", size=10, color="1F2937")
                    cell.alignment = Alignment(vertical="top")
                    cell.border = light_border

            for cell in worksheet[first_excel_row]:
                cell.border = Border(top=Side(style="medium", color="F2B134"))

        table = Table(displayName="MergedDataTable", ref=f"A1:{last_column}{last_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

        for column_index, column_name in enumerate(df.columns, start=1):
            sampled_values = df[column_name].head(1000)
            max_value_length = max(
                (len(str(value)) for value in sampled_values if not _is_blank(value)),
                default=0,
            )
            max_length = max(len(str(column_name)), max_value_length)
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(
                max(max_length + 2, 12),
                42,
            )

        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = False
        worksheet.auto_filter.ref = f"A1:{last_column}{last_row}"
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.print_title_rows = "1:1"
        worksheet.sheet_properties.tabColor = "087E8B"

        summary = workbook.create_sheet("Summary")
        summary.sheet_view.showGridLines = False
        summary.merge_cells("A1:F1")
        summary["A1"] = "Merge Summary"
        summary["A1"].fill = title_fill
        summary["A1"].font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
        summary["A1"].alignment = Alignment(vertical="center")
        summary.row_dimensions[1].height = 32

        metric_labels = (("A3", "Files merged"), ("C3", "Total rows"), ("E3", "Columns"))
        metric_values = (("B3", len(file_sources)), ("D3", len(df)), ("F3", len(df.columns)))
        for cell_address, label in metric_labels:
            summary[cell_address] = label
            summary[cell_address].fill = accent_fill
            summary[cell_address].font = Font(name="Aptos", bold=True, color="132238")
        for cell_address, value in metric_values:
            summary[cell_address] = value
            summary[cell_address].font = Font(name="Aptos", size=12, bold=True, color="132238")
            summary[cell_address].number_format = "#,##0"

        summary_headers = ("File #", "File name", "Rows merged", "Output data rows")
        for column_index, header in enumerate(summary_headers, start=1):
            cell = summary.cell(row=6, column=column_index, value=header)
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(vertical="center")

        for source_index, (filename, start_row, end_row) in enumerate(file_sources, start=1):
            output_row = source_index + 6
            row_fill = PatternFill("solid", fgColor=SOURCE_FILLS[(source_index - 1) % len(SOURCE_FILLS)])
            values = (
                source_index,
                filename,
                end_row - start_row + 1,
                f"{start_row + 2}:{end_row + 2}",
            )
            for column_index, value in enumerate(values, start=1):
                cell = summary.cell(row=output_row, column=column_index, value=value)
                cell.fill = row_fill
                cell.font = Font(name="Aptos", size=10, color="1F2937")
                cell.border = light_border
            summary.cell(row=output_row, column=3).number_format = "#,##0"

        summary_last_row = len(file_sources) + 6
        summary_table = Table(displayName="MergeSummaryTable", ref=f"A6:D{summary_last_row}")
        summary_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        summary.add_table(summary_table)
        summary.freeze_panes = "A7"
        summary.column_dimensions["A"].width = 10
        summary.column_dimensions["B"].width = 46
        summary.column_dimensions["C"].width = 16
        summary.column_dimensions["D"].width = 20
        summary.column_dimensions["E"].width = 14
        summary.column_dimensions["F"].width = 14
        summary.sheet_properties.tabColor = "F2B134"

        workbook.properties.title = "Merged File Output"
        workbook.properties.subject = "Combined Excel and CSV data"
        workbook.properties.creator = "Daily Report Tool"

    return output.getvalue()


def _file_signature(uploaded_files: Sequence) -> Tuple[Tuple[str, int], ...]:
    return tuple((uploaded_file.name, uploaded_file.size) for uploaded_file in uploaded_files)


def render_excel_merger_page():
    """Render the single-mode smart merger used by the sidebar route."""
    st.title("Merge Excel Files")
    st.caption("Combine 2 to 50 Excel/CSV files into one formatted Excel workbook.")
    st.markdown(
        "Matching headers are combined into one column, all rows are stacked, "
        "and the output contains one header row."
    )

    uploaded_files = st.file_uploader(
        "Choose Excel/CSV files (maximum 50)",
        type=["xlsx", "xls", "csv"],
        accept_multiple_files=True,
        key="smart_excel_merger_uploads",
    )

    if not uploaded_files:
        st.info("Select at least two files to begin.")
        return

    if len(uploaded_files) > MAX_MERGE_FILES:
        st.error(f"You selected {len(uploaded_files)} files. The maximum is {MAX_MERGE_FILES}.")
        return

    signature = _file_signature(uploaded_files)
    previous_result = st.session_state.get("smart_excel_merger_result")
    if previous_result and previous_result["signature"] != signature:
        st.session_state.pop("smart_excel_merger_result", None)
        previous_result = None

    file_list = pd.DataFrame(
        {
            "File": [uploaded_file.name for uploaded_file in uploaded_files],
            "Size": [f"{uploaded_file.size / (1024 * 1024):.2f} MB" for uploaded_file in uploaded_files],
        }
    )
    st.dataframe(file_list, use_container_width=True, hide_index=True)

    if len(uploaded_files) < 2:
        st.info("Add one more file to merge.")
        return

    if st.button(
        f"Merge {len(uploaded_files)} Files",
        type="primary",
        use_container_width=True,
    ):
        try:
            with st.spinner("Reading, cleaning, and merging files..."):
                merged_df, file_sources = merge_uploaded_files(uploaded_files)
                excel_bytes = generate_merged_excel(merged_df, file_sources)
                result = {
                    "signature": signature,
                    "bytes": excel_bytes,
                    "files": len(uploaded_files),
                    "rows": len(merged_df),
                    "columns": len(merged_df.columns),
                    "preview": merged_df.head(50),
                    "file_name": f"merged_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                }
                st.session_state.smart_excel_merger_result = result
                previous_result = result
        except Exception as error:
            st.session_state.pop("smart_excel_merger_result", None)
            st.error(f"Merge failed: {error}")
            return

    if previous_result:
        st.success("Files merged successfully. The Excel output is ready.")
        metric_files, metric_rows, metric_columns = st.columns(3)
        metric_files.metric("Files merged", previous_result["files"])
        metric_rows.metric("Rows", f"{previous_result['rows']:,}")
        metric_columns.metric("Columns", previous_result["columns"])

        with st.expander("Preview first 50 rows"):
            st.dataframe(previous_result["preview"], use_container_width=True, hide_index=True)

        st.download_button(
            "Download Formatted Excel",
            data=previous_result["bytes"],
            file_name=previous_result["file_name"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
