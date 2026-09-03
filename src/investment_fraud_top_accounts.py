"""Investment-fraud and digital-arrest Top 30 suspect account report.

The report joins qualifying complaints from the Additional Information report
to Layer 1 transactions in the Layerwise report, combines repeated numeric bank
accounts, and ranks the top 30 by total disputed amount.
"""

from __future__ import annotations

from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
import math
import re
import unicodedata
from typing import Any, Iterable

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.ui_styling import render_page_header_with_info


TOP30_CANDIDATE_LIMIT = 30
# Keep the existing public name for callers of this module.
TOP_ACCOUNT_LIMIT = TOP30_CANDIDATE_LIMIT

REPORT_HEADERS = [
    "Sr.No.",
    "Fraudster Bank Account Number",
    "All Acknowledgement Numbers",
    "ACK Count",
    "Bank Name",
    "IFSC Code",
    "Address",
    "District",
    "State",
    "Total Transactions",
    "Total Amount",
    "Total Disputed Amount",
]

ADDITIONAL_COLUMN_ALIASES = {
    "ack": [
        "Acknowledgement No.",
        "Acknowledgment No.",
        "Acknowledgement Number",
        "Acknowledgment Number",
        "ACK No.",
    ],
    "information": [
        "Crime Aditional Information",
        "Crime Additional Information",
        "Additional Information",
    ],
}

LAYERWISE_COLUMN_ALIASES = {
    "ack": [
        "Acknowledgement No.",
        "Acknowledgment No.",
        "Acknowledgement Number",
        "Acknowledgment Number",
        "ACK No.",
    ],
    "account": ["Account No.", "Account Number", "Bank Account Number"],
    "ifsc": ["IFSC Code", "IFSC"],
    "address": ["Address", "Bank Address"],
    "district": ["District", "Bank District"],
    "state": ["State", "Bank State"],
    "transaction_amount": ["Transaction Amount", "Amount"],
    "disputed_amount": ["Disputed Amount"],
    "bank": ["Bank/FIs", "Bank/FI", "Bank Name", "Bank"],
    "layer": ["Layers", "Layer"],
}


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    try:
        missing = pd.isna(value)
    except (TypeError, ValueError):
        missing = False
    if isinstance(missing, bool) and missing:
        return True
    return str(value).strip().lower() in {"", "nan", "none", "<na>", "nat"}


def _text_value(value: Any) -> str:
    return "" if _is_blank(value) else str(value).strip()


def _identifier_value(value: Any) -> str:
    """Return an Excel identifier without a trailing .0 or scientific notation."""
    if _is_blank(value):
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            return ""
        if value.is_integer():
            return str(int(value))
        return format(value, "f").rstrip("0").rstrip(".")

    text = str(value).strip()
    if "e" not in text.lower():
        return text
    try:
        number = Decimal(text)
    except InvalidOperation:
        return text
    if number == number.to_integral_value():
        return format(number.quantize(Decimal("1")), "f")
    return format(number.normalize(), "f").rstrip("0").rstrip(".")


def _normalized(value: Any) -> str:
    text = unicodedata.normalize("NFKD", _text_value(value))
    text = "".join(character for character in text if not unicodedata.combining(character))
    return re.sub(r"[^A-Z0-9]+", " ", text.upper()).strip()


def _header_key(value: Any) -> str:
    return _normalized(value).replace(" ", "")


def _edit_distance(left: str, right: str) -> int:
    previous = list(range(len(right) + 1))
    for row, left_character in enumerate(left, start=1):
        current = [row]
        for column, right_character in enumerate(right, start=1):
            current.append(
                min(
                    current[column - 1] + 1,
                    previous[column] + 1,
                    previous[column - 1] + (left_character != right_character),
                )
            )
        previous = current
    return previous[-1]


def is_investment_fraud(value: Any) -> bool:
    """Match nearby investment/fraud words while tolerating common misspellings."""
    words = [word for word in _normalized(value).split() if word]
    investment_positions = []
    fraud_positions = []

    for index, word in enumerate(words):
        if word.startswith("INVEST") or _edit_distance(word, "INVESTMENT") <= 2:
            investment_positions.append(index)
        if word.startswith("FRAUD") or _edit_distance(word, "FRAUD") <= 1:
            fraud_positions.append(index)

    return any(
        abs(investment_index - fraud_index) <= 3
        for investment_index in investment_positions
        for fraud_index in fraud_positions
    )


def is_digital_arrest(value: Any) -> bool:
    """Match nearby digital/arrest words while tolerating common misspellings."""
    words = [word for word in _normalized(value).split() if word]
    digital_positions = [
        index
        for index, word in enumerate(words)
        if word.startswith("DIGIT") or _edit_distance(word, "DIGITAL") <= 2
    ]
    arrest_positions = [
        index
        for index, word in enumerate(words)
        if word.startswith("ARREST") or _edit_distance(word, "ARREST") <= 2
    ]
    return any(
        abs(digital_index - arrest_index) <= 3
        for digital_index in digital_positions
        for arrest_index in arrest_positions
    )


def is_qualifying_fraud(value: Any) -> bool:
    """Return whether the complaint belongs in the combined Top 30 report."""
    return is_investment_fraud(value) or is_digital_arrest(value)


def _find_column(columns: Iterable[Any], aliases: Iterable[str], label: str) -> Any:
    column_list = list(columns)
    keyed_columns = [(_header_key(column), column) for column in column_list]
    alias_keys = [_header_key(alias) for alias in aliases]

    for key, column in keyed_columns:
        if key and key in alias_keys:
            return column

    for key, column in keyed_columns:
        if key and any(key in alias or alias in key for alias in alias_keys):
            return column

    available = ", ".join(_text_value(column) for column in column_list)
    raise ValueError(f"Required column not found: {label}. Available columns: {available}")


def _numeric_value(value: Any) -> float:
    if _is_blank(value):
        return 0.0
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value) if math.isfinite(float(value)) else 0.0

    text = str(value).strip()
    negative = text.startswith("(") and text.endswith(")")
    cleaned = re.sub(r"[₹,\s]", "", text.strip("()"))
    try:
        number = float(cleaned)
    except ValueError:
        return 0.0
    return -number if negative else number


def _is_layer_one(value: Any) -> bool:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            return float(value) == 1
        except (TypeError, ValueError):
            return False
    return _normalized(value).replace(" ", "") in {"1", "LAYER1"}


def _is_bank_name(value: Any) -> bool:
    return "BANK" in _normalized(value)


def _first_non_blank(existing: str, candidate: Any) -> str:
    return existing or _text_value(candidate)


def build_investment_fraud_report(
    additional_df: pd.DataFrame,
    layerwise_df: pd.DataFrame,
    limit: int = TOP_ACCOUNT_LIMIT,
) -> tuple[pd.DataFrame, dict[str, int]]:
    """Build the ranked account table and processing summary."""
    if additional_df.empty:
        raise ValueError("The Additional Information report does not contain data rows.")
    if layerwise_df.empty:
        raise ValueError("The Layerwise report does not contain data rows.")

    additional_columns = {
        key: _find_column(additional_df.columns, aliases, label)
        for key, aliases, label in [
            ("ack", ADDITIONAL_COLUMN_ALIASES["ack"], "Acknowledgement No."),
            (
                "information",
                ADDITIONAL_COLUMN_ALIASES["information"],
                "Crime Additional Information",
            ),
        ]
    }
    layerwise_labels = {
        "ack": "Acknowledgement No.",
        "account": "Account No.",
        "ifsc": "IFSC Code",
        "address": "Address",
        "district": "District",
        "state": "State",
        "transaction_amount": "Transaction Amount",
        "disputed_amount": "Disputed Amount",
        "bank": "Bank/FIs",
        "layer": "Layers",
    }
    layerwise_columns = {
        key: _find_column(layerwise_df.columns, aliases, layerwise_labels[key])
        for key, aliases in LAYERWISE_COLUMN_ALIASES.items()
    }

    qualifying_acknowledgements = {
        _identifier_value(row[additional_columns["ack"]])
        for _, row in additional_df.iterrows()
        if is_qualifying_fraud(row[additional_columns["information"]])
        and _identifier_value(row[additional_columns["ack"]])
    }

    matched_rows = []
    for _, row in layerwise_df.iterrows():
        acknowledgement = _identifier_value(row[layerwise_columns["ack"]])
        if acknowledgement in qualifying_acknowledgements and _is_layer_one(
            row[layerwise_columns["layer"]]
        ):
            matched_rows.append(row)

    accounts: dict[str, dict[str, Any]] = {}
    for row in matched_rows:
        account_number = _identifier_value(row[layerwise_columns["account"]])
        if not re.fullmatch(r"[0-9]+", account_number):
            continue

        account_key = account_number.upper()
        account = accounts.setdefault(
            account_key,
            {
                "account_number": account_number,
                "acknowledgements": set(),
                "bank_name": "",
                "ifsc_code": "",
                "address": "",
                "district": "",
                "state": "",
                "total_transactions": 0,
                "total_amount": 0.0,
                "total_disputed_amount": 0.0,
            },
        )
        acknowledgement = _identifier_value(row[layerwise_columns["ack"]])
        if acknowledgement:
            account["acknowledgements"].add(acknowledgement)
        account["bank_name"] = _first_non_blank(
            account["bank_name"], row[layerwise_columns["bank"]]
        )
        account["ifsc_code"] = _first_non_blank(
            account["ifsc_code"], row[layerwise_columns["ifsc"]]
        )
        account["address"] = _first_non_blank(
            account["address"], row[layerwise_columns["address"]]
        )
        account["district"] = _first_non_blank(
            account["district"], row[layerwise_columns["district"]]
        )
        account["state"] = _first_non_blank(
            account["state"], row[layerwise_columns["state"]]
        )
        account["total_transactions"] += 1
        account["total_amount"] += _numeric_value(
            row[layerwise_columns["transaction_amount"]]
        )
        account["total_disputed_amount"] += _numeric_value(
            row[layerwise_columns["disputed_amount"]]
        )

    bank_accounts = [
        account for account in accounts.values() if _is_bank_name(account["bank_name"])
    ]
    ranked_accounts = sorted(
        bank_accounts,
        key=lambda account: (
            -account["total_disputed_amount"],
            -account["total_amount"],
            account["account_number"].upper(),
        ),
    )[: max(0, limit)]

    rows = []
    for rank, account in enumerate(ranked_accounts, start=1):
        acknowledgements = sorted(account["acknowledgements"])
        rows.append(
            {
                "Sr.No.": rank,
                "Fraudster Bank Account Number": account["account_number"],
                "All Acknowledgement Numbers": ";".join(acknowledgements),
                "ACK Count": len(acknowledgements),
                "Bank Name": account["bank_name"],
                "IFSC Code": account["ifsc_code"],
                "Address": account["address"],
                "District": account["district"],
                "State": account["state"],
                "Total Transactions": account["total_transactions"],
                "Total Amount": account["total_amount"],
                "Total Disputed Amount": account["total_disputed_amount"],
            }
        )

    report_df = pd.DataFrame(rows, columns=REPORT_HEADERS)
    summary = {
        "qualifying_acknowledgements": len(qualifying_acknowledgements),
        "matched_layer_one_transactions": len(matched_rows),
        "matched_accounts": len(bank_accounts),
        "excluded_non_bank_accounts": len(accounts) - len(bank_accounts),
        "output_accounts": len(ranked_accounts),
    }
    return report_df, summary


def generate_investment_fraud_excel(
    report_df: pd.DataFrame,
    report_date: date | datetime | None = None,
) -> tuple[bytes, str]:
    """Create the formatted Excel report and return its bytes and filename."""
    if report_date is None:
        report_date = datetime.now() - timedelta(days=1)
    date_label = report_date.strftime("%d-%m-%Y")
    title = (
        f"{date_label} Investment Fraud and Digital Arrest Top 30 "
        "Suspect Accounts from Layer 1"
    )
    filename = f"{title}.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Top 30 Suspect Accounts"
    worksheet.sheet_view.showGridLines = False

    navy = "17365D"
    blue = "2F75B5"
    pale_blue = "DDEBF7"
    pale_border = "D9E2F3"
    body_text = "203040"
    thin_border = Border(
        left=Side(style="thin", color=pale_border),
        right=Side(style="thin", color=pale_border),
        top=Side(style="thin", color=pale_border),
        bottom=Side(style="thin", color=pale_border),
    )

    worksheet.merge_cells("A1:L1")
    worksheet["A1"] = title
    worksheet["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    worksheet["A1"].fill = PatternFill("solid", fgColor=navy)
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet["A1"].border = Border(
        left=Side(style="medium", color=navy),
        right=Side(style="medium", color=navy),
        top=Side(style="medium", color=navy),
        bottom=Side(style="medium", color=navy),
    )
    worksheet.row_dimensions[1].height = 34

    worksheet.merge_cells("A2:L2")
    worksheet["A2"] = (
        f"{len(report_df)} verified bank accounts • Layer 1 • "
        "Ranked by total disputed amount"
    )
    worksheet["A2"].font = Font(
        name="Calibri", size=10, bold=True, italic=True, color="1F4E78"
    )
    worksheet["A2"].fill = PatternFill("solid", fgColor=pale_blue)
    worksheet["A2"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet["A2"].border = Border(
        left=Side(style="thin", color="9EADBA"),
        right=Side(style="thin", color="9EADBA"),
        top=Side(style="thin", color="9EADBA"),
        bottom=Side(style="thin", color="9EADBA"),
    )
    worksheet.row_dimensions[2].height = 22

    for column_index, header in enumerate(REPORT_HEADERS, start=1):
        cell = worksheet.cell(row=3, column=column_index, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = thin_border
    worksheet.row_dimensions[3].height = 32

    medal_fills = {4: "FFF2CC", 5: "E7E6E6", 6: "FCE4D6"}
    for row_index, row in enumerate(report_df.itertuples(index=False, name=None), start=4):
        row_fill = medal_fills.get(row_index)
        if row_fill is None and row_index % 2 == 1:
            row_fill = "F4F8FC"

        for column_index, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            cell.font = Font(
                name="Calibri",
                size=10,
                bold=row_index in medal_fills,
                color=body_text,
            )
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if row_fill:
                cell.fill = PatternFill("solid", fgColor=row_fill)

        worksheet.cell(row=row_index, column=1).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        for column_index in (2, 3, 5, 6, 7, 8, 9):
            worksheet.cell(row=row_index, column=column_index).alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=column_index in (5, 7, 8, 9),
            )
        for column_index in (2, 3, 6):
            identifier_cell = worksheet.cell(row=row_index, column=column_index)
            identifier_cell.number_format = "@"
            identifier_cell.quotePrefix = True
        for column_index in (4, 10):
            worksheet.cell(row=row_index, column=column_index).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            worksheet.cell(row=row_index, column=column_index).number_format = "#,##0"
        for column_index in (11, 12):
            worksheet.cell(row=row_index, column=column_index).alignment = Alignment(
                horizontal="right", vertical="center"
            )
            worksheet.cell(row=row_index, column=column_index).number_format = (
                '"₹"#,##0.00'
            )
        worksheet.row_dimensions[row_index].height = 24

    widths = [8, 25, 34, 11, 36, 17, 46, 20, 20, 18, 19, 21]
    for column_index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    last_row = max(3, len(report_df) + 3)
    worksheet.auto_filter.ref = f"A3:L{last_row}"
    worksheet.freeze_panes = "A4"
    worksheet.print_title_rows = "1:3"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.print_area = f"A1:L{last_row}"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue(), filename


def _read_excel_upload(uploaded_file: Any) -> pd.DataFrame:
    return pd.read_excel(BytesIO(uploaded_file.getvalue()), dtype=object)


def render_investment_fraud_top_accounts_page() -> None:
    """Render the Investment Fraud and Digital Arrest Top 30 page."""
    render_page_header_with_info("investment_fraud_top_accounts")

    st.markdown(
        "Upload the Additional Information and Layerwise reports. The page "
        "finds investment-fraud and digital-arrest complaints, matches their "
        "Layer 1 transactions, combines repeated bank accounts, and ranks the top 30."
    )
    st.markdown("---")

    additional_column, layerwise_column = st.columns(2)
    with additional_column:
        st.subheader("Additional Information Report")
        additional_file = st.file_uploader(
            "Upload Additional Information Excel file",
            type=["xlsx", "xls"],
            key="investment_fraud_additional_upload",
            help="The file must include Acknowledgement No. and Crime Additional Information.",
        )
    with layerwise_column:
        st.subheader("Layerwise Report")
        layerwise_file = st.file_uploader(
            "Upload Layerwise Excel file",
            type=["xlsx", "xls"],
            key="investment_fraud_layerwise_upload",
            help="The file must include account, bank, layer, amount, and location columns.",
        )

    with st.expander("Report rules", expanded=False):
        st.markdown(
            """
            - Detects common spelling variations of **investment fraud** and **digital arrest**.
            - Matches acknowledgement numbers between the two reports.
            - Keeps **Layer 1** transactions and entries whose bank name contains **Bank**.
            - Keeps only account numbers made entirely of ASCII digits (0-9).
            - Combines repeated account numbers and sums Transaction Amount and Disputed Amount.
            - Ranks up to 30 accounts by Total Disputed Amount, then Total Amount.
            """
        )

    if not additional_file or not layerwise_file:
        st.info("Upload both Excel files to generate the report.")
        return

    try:
        with st.spinner("Matching complaints and ranking Layer 1 bank accounts..."):
            additional_df = _read_excel_upload(additional_file)
            layerwise_df = _read_excel_upload(layerwise_file)
            report_df, summary = build_investment_fraud_report(
                additional_df, layerwise_df
            )
            excel_bytes, filename = generate_investment_fraud_excel(report_df)
    except (ValueError, KeyError, OSError) as error:
        st.error(str(error))
        return
    except Exception as error:
        st.error(f"Unable to generate the report: {error}")
        return

    metric_columns = st.columns(4)
    metric_columns[0].metric(
        "Qualifying ACKs", f"{summary['qualifying_acknowledgements']:,}"
    )
    metric_columns[1].metric(
        "Matched Layer 1 rows", f"{summary['matched_layer_one_transactions']:,}"
    )
    metric_columns[2].metric(
        "Eligible bank accounts", f"{summary['matched_accounts']:,}"
    )
    metric_columns[3].metric("Accounts in report", f"{summary['output_accounts']:,}")

    if summary["excluded_non_bank_accounts"]:
        st.caption(
            f"{summary['excluded_non_bank_accounts']:,} account(s) were excluded "
            "because the bank name did not contain “Bank”."
        )

    st.subheader("Ranked account preview")
    if report_df.empty:
        st.warning("No qualifying Layer 1 bank accounts were found.")
    else:
        preview_df = report_df.copy()
        preview_df["Total Amount"] = preview_df["Total Amount"].map(
            lambda value: f"₹{value:,.2f}"
        )
        preview_df["Total Disputed Amount"] = preview_df[
            "Total Disputed Amount"
        ].map(lambda value: f"₹{value:,.2f}")
        st.dataframe(preview_df, use_container_width=True, hide_index=True)

    st.download_button(
        label="Download Investment Fraud & Digital Arrest Top 30 Excel Report",
        data=excel_bytes,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )
