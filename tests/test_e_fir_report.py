import io
from datetime import date, datetime

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from src.e_fir_report import (
    REPORT_COLUMNS,
    build_efir_report,
    generate_efir_excel,
    process_efir_report_files,
)


def _zero_efir_workbook_bytes() -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "TotalZeroEFir"
    summary.append(["State/UTs", "Total Complaints"])
    summary.append(["GUJARAT", 3])

    details = workbook.create_sheet("Details")
    details.append(
        [
            "S No.",
            "Acknowledgement No",
            "Total Reported Amount",
            "e-Zero FIR No.",
            "Date of e-Zero FIR",
        ]
    )
    details.append(
        [1, "ACK002", 5000, "011200000000001", datetime(2026, 8, 30, 9, 0)]
    )
    details.append(
        [2, "ACK001", 12500, "011200000000002", datetime(2026, 8, 30, 8, 0)]
    )
    details.append(
        [3, "ACK999", 9999, "011200000000999", datetime(2026, 8, 29, 23, 0)]
    )

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_build_efir_report_uses_efir_amount_and_sorts_by_efir_time():
    crime_df = pd.DataFrame(
        {
            "Acknowledgement No.": ["ACK001", "ACK001", "ACK002", "ACK003"],
            "Name of Complainant": ["Asha", "Asha", "Bharat", "Chirag"],
            "Amount": ["1,000", 250.5, 500, 900],
            "District ": ["SURAT", "SURAT", "RAJKOT", "VADODARA"],
            "Police Station": ["VESU", "VESU", "A-DIVISION", "B-DIVISION"],
        }
    )
    efir_df = pd.DataFrame(
        {
            "Acknowledgement No": ["ACK002", "ACK001", "ACK999"],
            "e-Zero FIR No.": [
                "011200000000001",
                "011200000000002",
                "011200000000999",
            ],
            "Total Reported Amount": [5000, 12500, 9999],
            "Date of e-Zero FIR": [
                datetime(2026, 8, 30, 9, 0),
                datetime(2026, 8, 30, 8, 0),
                datetime(2026, 8, 29, 23, 0),
            ],
        }
    )

    report, summary = build_efir_report(crime_df, efir_df, date(2026, 8, 30))

    assert list(report.columns) == REPORT_COLUMNS
    assert report["ACK No."].tolist() == ["ACK001", "ACK002"]
    assert report["Fraudulent Amount"].tolist() == [12500.0, 5000.0]
    assert report.loc[0, "Name"] == "Asha"
    assert report.loc[0, "District"] == "SURAT"
    assert report.loc[0, "Police Station"] == "VESU"
    assert report.loc[0, "E-FIR No."] == "011200000000002"
    assert summary["matched_acknowledgements"] == 2
    assert summary["combined_crime_rows"] == 1
    assert summary["unmatched_crime_acknowledgements"] == 1
    assert summary["unmatched_efir_acknowledgements"] == 0


def test_process_efir_report_files_uses_details_sheet():
    crime_csv = (
        "Acknowledgement No.,Name of Complainant,Amount,District,Police Station\n"
        "ACK001,Asha,100,SURAT,VESU\n"
        "ACK001,Asha,200,SURAT,VESU\n"
        "ACK002,Bharat,500,RAJKOT,A-DIVISION\n"
    ).encode("utf-8")

    result = process_efir_report_files(
        crime_csv,
        "Crime_Report.csv",
        _zero_efir_workbook_bytes(),
        "Zero_E_FIR.xlsx",
        date(2026, 8, 30),
    )

    assert result.report["ACK No."].tolist() == ["ACK001", "ACK002"]
    assert result.report["Fraudulent Amount"].tolist() == [12500.0, 5000.0]
    assert result.filename.startswith("E_FIR_Report_")
    assert result.filename.endswith(".xlsx")


def test_generate_efir_excel_preserves_identifiers_and_amount_format():
    report = pd.DataFrame(
        [
            {
                "Sr.": 1,
                "ACK No.": "31108260204908",
                "Name": "MANOJKUMAR TRIVEDI",
                "Fraudulent Amount": 95000.5,
                "District": "BANASKANTHA",
                "Police Station": "DEESA SOUTH",
                "E-FIR No.": "011195004260192",
            }
        ],
        columns=REPORT_COLUMNS,
    )

    workbook = load_workbook(io.BytesIO(generate_efir_excel(report)), data_only=True)
    try:
        worksheet = workbook["E-FIR Report"]
        assert [cell.value for cell in worksheet[1]] == REPORT_COLUMNS
        assert worksheet["B2"].value == "31108260204908"
        assert worksheet["B2"].number_format == "@"
        assert worksheet["D2"].value == 95000.5
        assert worksheet["D2"].number_format == "#,##0.00"
        assert worksheet["G2"].value == "011195004260192"
        assert worksheet["G2"].number_format == "@"
        assert worksheet.freeze_panes == "A2"
    finally:
        workbook.close()


def test_generate_efir_excel_handles_blank_amount():
    report = pd.DataFrame(
        [[1, "ACK001", "Asha", float("nan"), "SURAT", "VESU", "EFIR001"]],
        columns=REPORT_COLUMNS,
    )

    workbook = load_workbook(io.BytesIO(generate_efir_excel(report)), data_only=True)
    try:
        assert workbook["E-FIR Report"]["D2"].value is None
    finally:
        workbook.close()


def test_build_efir_report_reports_missing_required_columns():
    crime_df = pd.DataFrame({"Acknowledgement No.": ["ACK001"]})
    efir_df = pd.DataFrame(
        {
            "Acknowledgement No": ["ACK001"],
            "e-Zero FIR No.": ["EFIR001"],
            "Total Reported Amount": [1000],
            "Date of e-Zero FIR": [datetime(2026, 8, 30, 8, 0)],
        }
    )

    with pytest.raises(ValueError, match="Crime Report is missing required column"):
        build_efir_report(crime_df, efir_df, date(2026, 8, 30))


def test_build_efir_report_uses_selected_date_as_master_filter():
    crime_df = pd.DataFrame(
        {
            "Acknowledgement No.": ["ACK001", "ACK002"],
            "Name of Complainant": ["Asha", "Bharat"],
            "Amount": [100, 200],
            "District": ["SURAT", "RAJKOT"],
            "Police Station": ["VESU", "A-DIVISION"],
        }
    )
    efir_df = pd.DataFrame(
        {
            "Acknowledgement No": ["ACK001", "ACK002"],
            "e-Zero FIR No.": ["EFIR001", "EFIR002"],
            "Total Reported Amount": [1000, 2000],
            "Date of e-Zero FIR": [
                datetime(2026, 8, 29, 23, 59),
                datetime(2026, 8, 30, 0, 1),
            ],
        }
    )

    report, summary = build_efir_report(crime_df, efir_df, date(2026, 8, 30))

    assert report["ACK No."].tolist() == ["ACK002"]
    assert report["Fraudulent Amount"].tolist() == [2000.0]
    assert summary["efir_acknowledgements"] == 1
    assert summary["matched_acknowledgements"] == 1
